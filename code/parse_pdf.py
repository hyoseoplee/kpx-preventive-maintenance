#!/usr/bin/env python3
"""
Parse KPX weekly maintenance schedule PDFs and generate a merged CSV.

Usage:
    python parse_pdf.py <year>
    python parse_pdf.py 2022
    python parse_pdf.py 2025

Input:  data/kpx-data/{year}/*.pdf
Output: results/{year}/maintenance_schedule.csv

Columns: 회원사, 연료원, 발전기명, 설비용량, 시작일, 시작시간, 종료일, 종료시간, 광역지역, 세부지역
"""

import sys
import os
import re
import csv
import glob
from datetime import datetime
from collections import defaultdict

import pdfplumber
import openpyxl

# Resolve repo root (one level up from code/)
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

FUEL_MAP = {
    '유연탄': 'Coal',
    '무연탄': 'Coal',
    '유류': 'Oil',
    '원자력': 'Nuclear',
}

# PDF 회원사 → Excel 회원사 이름 매핑 (약칭 사용하는 경우)
COMPANY_NORMALIZE = {
    '한국남동발전(주)': '남동',
    '한국남부발전(주)': '남부',
    '한국동서발전(주)': '동서',
    '한국서부발전(주)': '서부',
    '한국중부발전(주)': '중부',
    '한국수력원자력(주)': '한수원',
}


# ── Excel master data ──────────────────────────────────────────────────────

def load_gen_master(xlsx_path):
    """
    Load generator master data from Excel.
    Returns dict: (회원사, base_name) → (광역지역, 세부지역)
    Also returns a fallback dict: base_name → (광역지역, 세부지역) for cases
    where company matching fails.
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    master_by_company = {}  # (회원사, base_name) → (광역지역, 세부지역)
    master_by_name = {}     # base_name → (광역지역, 세부지역)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        company = str(row[0]).strip() if row[0] else ''
        gen_name = str(row[1]).strip() if row[1] else ''
        region = str(row[9]).strip() if row[9] else ''
        detail = str(row[10]).strip() if row[10] else ''

        base = extract_base_name(gen_name)
        if base:
            key = (company, base)
            if key not in master_by_company:
                master_by_company[key] = (region, detail)
            if base not in master_by_name:
                master_by_name[base] = (region, detail)

    return master_by_company, master_by_name


def extract_base_name(name):
    """
    Extract the base plant name from various naming conventions.
    '분당복합 CC1 GT1' → '분당'
    '영흥#3' → '영흥'
    'GS당진복합 CC1 GT1' → 'GS당진'
    '신보령화력#1' → '신보령'
    """
    name = name.strip()
    name = re.sub(r'(복합|화력|열병합|그린파워|천연가스)\b.*', '', name)
    name = re.sub(r'\s*(CC\d+\s*)?(GT|ST|#)\d*.*', '', name)
    name = re.sub(r'\s+\d+호기.*', '', name)
    return name.strip()


def find_location(company, gen_name, master_by_company, master_by_name):
    """
    Match PDF generator to Excel master data for location info.
    First tries matching by (회원사, base_name), then falls back to base_name only.
    """
    base = extract_base_name(gen_name)

    # Normalize PDF company name to Excel convention
    norm_company = COMPANY_NORMALIZE.get(company, company)

    # Try exact (company, base) match
    key = (norm_company, base)
    if key in master_by_company:
        return master_by_company[key]

    # Try all Excel companies with the same base name
    for (comp, b), loc in master_by_company.items():
        if b == base:
            return loc

    # Fallback: base name only
    if base in master_by_name:
        return master_by_name[base]

    # Try without common prefixes
    for prefix in ['신', 'GS']:
        if base.startswith(prefix) and base[len(prefix):] in master_by_name:
            return master_by_name[base[len(prefix):]]

    # Try partial match
    for key in master_by_name:
        if base and key.startswith(base):
            return master_by_name[key]

    return ('', '')


# ── PDF extraction ─────────────────────────────────────────────────────────

def parse_datetime(date_str, time_str):
    date_str = date_str.strip()
    time_str = time_str.strip()
    if re.match(r'^\d:', time_str):
        time_str = '0' + time_str
    return datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")


def extract_table_from_pdf(pdf_path):
    rows = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if row is None or len(row) < 8:
                        continue
                    if row[0] and ('회원사' in str(row[0]) or '시작일' in str(row[0])):
                        continue
                    if row[0] and ('주 간' in str(row[0]) or '대상기간' in str(row[0])
                                   or '수급계획' in str(row[0])):
                        continue

                    try:
                        company = str(row[0]).strip() if row[0] else ''
                        fuel_type = str(row[1]).strip() if row[1] else None
                        generator = str(row[2]).strip() if row[2] else None
                        capacity_str = str(row[3]).strip().replace(',', '') if row[3] else None
                        start_date = str(row[4]).strip() if row[4] else None
                        start_time = str(row[5]).strip() if row[5] else None
                        end_date = str(row[6]).strip() if row[6] else None
                        end_time = str(row[7]).strip() if row[7] else None

                        if not generator or not start_date or not end_date:
                            continue
                        if not re.match(r'\d{4}-\d{2}-\d{2}', start_date):
                            continue
                        if not re.match(r'\d{4}-\d{2}-\d{2}', end_date):
                            continue

                        capacity = float(capacity_str)
                        start_dt = parse_datetime(start_date, start_time)
                        end_dt = parse_datetime(end_date, end_time)
                        fuel_mapped = FUEL_MAP.get(fuel_type, fuel_type)

                        rows.append({
                            '회원사': company,
                            '연료원': fuel_mapped,
                            '발전기명': generator,
                            '설비용량': capacity,
                            '시작일': start_date,
                            '시작시간': start_time,
                            '종료일': end_date,
                            '종료시간': end_time,
                            '_start_dt': start_dt,
                            '_end_dt': end_dt,
                        })
                    except (ValueError, TypeError, IndexError):
                        continue
    return rows


def merge_overlapping_periods(records):
    groups = defaultdict(list)
    for r in records:
        key = (r['회원사'], r['연료원'], r['발전기명'], r['설비용량'])
        groups[key].append(r)

    merged = []
    for (company, fuel, gen, cap), entries in groups.items():
        entries.sort(key=lambda x: x['_start_dt'])
        merged_intervals = []
        for e in entries:
            if merged_intervals and e['_start_dt'] <= merged_intervals[-1]['_end_dt']:
                if e['_end_dt'] > merged_intervals[-1]['_end_dt']:
                    merged_intervals[-1]['_end_dt'] = e['_end_dt']
                    merged_intervals[-1]['종료일'] = e['종료일']
                    merged_intervals[-1]['종료시간'] = e['종료시간']
            else:
                merged_intervals.append(dict(e))
        merged.extend(merged_intervals)
    return merged


# ── Main ───────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python parse_pdf.py <year>")
        print("Example: python parse_pdf.py 2022")
        sys.exit(1)

    year = sys.argv[1]
    pdf_dir = os.path.join(REPO_ROOT, 'data', 'kpx-data', year)
    output_dir = os.path.join(REPO_ROOT, 'results', year)
    os.makedirs(output_dir, exist_ok=True)

    if not os.path.isdir(pdf_dir):
        print(f"Error: PDF directory not found: {pdf_dir}")
        sys.exit(1)

    # Load Excel master data
    xlsx_files = glob.glob(os.path.join(REPO_ROOT, 'data', 'HOME_발전설비_발전기별_발전기현황-*.xlsx'))
    master_by_company = {}
    master_by_name = {}
    if xlsx_files:
        xlsx_path = sorted(xlsx_files)[-1]
        print(f"Loading master data: {os.path.basename(xlsx_path)}")
        master_by_company, master_by_name = load_gen_master(xlsx_path)
        print(f"  {len(master_by_company)} (company, plant) entries, {len(master_by_name)} base names")
    else:
        print("Warning: No Excel master file found, skipping location data")

    # Extract from PDFs
    pdf_files = sorted(glob.glob(os.path.join(pdf_dir, "*.pdf")))
    print(f"\nFound {len(pdf_files)} PDFs in {pdf_dir}")

    all_rows = []
    for i, pdf_path in enumerate(pdf_files):
        fname = os.path.basename(pdf_path)
        rows = extract_table_from_pdf(pdf_path)
        print(f"  [{i+1}/{len(pdf_files)}] {fname}: {len(rows)} rows")
        all_rows.extend(rows)

    print(f"\nTotal raw rows: {len(all_rows)}")

    # Deduplicate
    seen = set()
    unique_rows = []
    for r in all_rows:
        key = (r['회원사'], r['연료원'], r['발전기명'], r['설비용량'],
               r['시작일'], r['시작시간'], r['종료일'], r['종료시간'])
        if key not in seen:
            seen.add(key)
            unique_rows.append(r)
    print(f"After dedup: {len(unique_rows)} rows")

    # Merge overlapping periods
    merged = merge_overlapping_periods(unique_rows)
    print(f"After merging: {len(merged)} rows")

    # Sort by generator name, then start date
    merged.sort(key=lambda x: (x['발전기명'], x['시작일']))

    # Write CSV with location columns
    output_path = os.path.join(output_dir, 'maintenance_schedule.csv')
    fieldnames = ['회원사', '연료원', '발전기명', '설비용량', '시작일', '시작시간', '종료일', '종료시간',
                  '광역지역', '세부지역']

    with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in merged:
            region, detail = find_location(
                r['회원사'], r['발전기명'], master_by_company, master_by_name
            ) if master_by_company else ('', '')
            writer.writerow({
                '회원사': r['회원사'],
                '연료원': r['연료원'],
                '발전기명': r['발전기명'],
                '설비용량': r['설비용량'],
                '시작일': r['시작일'],
                '시작시간': r['시작시간'],
                '종료일': r['종료일'],
                '종료시간': r['종료시간'],
                '광역지역': region,
                '세부지역': detail,
            })

    print(f"\nSaved to: {output_path}")

    # Summary
    fuel_counts = defaultdict(int)
    for r in merged:
        fuel_counts[r['연료원']] += 1
    print(f"Fuel type distribution: {dict(sorted(fuel_counts.items(), key=lambda x: -x[1]))}")


if __name__ == "__main__":
    main()
